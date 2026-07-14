#!/usr/bin/env python3
"""
Generate L30 Descriptions using IBM watsonx.ai (Llama 3)
---------------------------------------------------------
Reads the UT Flat Excel file, finds rows with a blank L30 Description,
calls watsonx.ai Llama 3 to generate a 2-3 sentence description for each,
and writes the result back to the Excel file after every row (safe to re-run).

Setup:
    pip install ibm-watsonx-ai openpyxl

Required environment variables (or edit the CONFIG section below):
    WATSONX_API_KEY   - your IBM Cloud API key
    WATSONX_PROJECT_ID - your watsonx.ai project ID
    WATSONX_URL       - watsonx.ai endpoint, e.g. https://us-south.ml.cloud.ibm.com

Usage:
    python scripts/generate_l30_descriptions.py
    python scripts/generate_l30_descriptions.py --software-only   # skip IBM Infrastructure rows
    python scripts/generate_l30_descriptions.py --dry-run         # print prompts, no API calls
"""

import os
import sys
import time
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install ibm-watsonx-ai openpyxl")

try:
    from ibm_watsonx_ai import APIClient, Credentials
    from ibm_watsonx_ai.foundation_models import ModelInference
    from ibm_watsonx_ai.foundation_models.schema import TextChatParameters
except ImportError:
    sys.exit("Missing dependency: pip install ibm-watsonx-ai openpyxl")

# ── CONFIG ────────────────────────────────────────────────────────────────────
EXCEL_PATH = Path("../web/UT Flat - Effective on 2026-07-06.xlsx")
SHEET_NAME = "Flat UT"

WATSONX_URL        = os.getenv("WATSONX_URL", "https://eu-gb.ml.cloud.ibm.com")
WATSONX_API_KEY    = os.getenv("WATSONX_API_KEY", "")
WATSONX_PROJECT_ID = os.getenv("WATSONX_PROJECT_ID", "")

MODEL_ID = "meta-llama/llama-4-maverick-17b-128e-instruct-fp8"

# Column indices (1-based for openpyxl)
COL_L10  = 1
COL_L15  = 2
COL_L17  = 3
COL_L20  = 4
COL_L30  = 5
COL_DESC = 6

# L30 values that are too vague/internal to describe — leave blank
SKIP_VALUES = {
    "Bob", "Project Willow - Data", "Calistoga", "Janes", "Open Core",
    "SaaS Committed Spend", "HashiCorp Committed Spend",
    "Data Fabric Ecosystem & Emerging", "IT Automation Heritage",
    "Sovereign Core", "Restricted Use Only - zExpansion Future Opportunities",
    "Misc SW related PIDs", "Rental Revenue", "Printers",
    "Bob Premium Package for i",
}
# ─────────────────────────────────────────────────────────────────────────────


def build_prompt(l10: str, l15: str, l17: str, l20: str, l30: str) -> str:
    return (
        f"You are a technical writer creating brief product descriptions for an IBM software catalog.\n\n"
        f"Write a 2-3 sentence description for the following IBM product or service. "
        f"Be factual, clear, and concise. Do not start with the product name as the first word. "
        f"Do not use bullet points or headers. Output only the description, nothing else.\n\n"
        f"Portfolio context: {l10} > {l15} > {l17} > {l20}\n"
        f"Product name: {l30}\n\n"
        f"Description:"
    )


def get_model(dry_run: bool):
    if dry_run:
        return None
    if not WATSONX_API_KEY:
        sys.exit("Error: WATSONX_API_KEY is not set.")
    if not WATSONX_PROJECT_ID:
        sys.exit("Error: WATSONX_PROJECT_ID is not set.")

    credentials = Credentials(url=WATSONX_URL, api_key=WATSONX_API_KEY)
    client = APIClient(credentials)
    model = ModelInference(
        model_id=MODEL_ID,
        api_client=client,
        project_id=WATSONX_PROJECT_ID,
        params=TextChatParameters(
            max_tokens=150,
            temperature=0.3,
        ),
    )
    return model


def generate_description(model, prompt: str) -> str:
    messages = [{"role": "user", "content": prompt}]
    response = model.chat(messages=messages)
    return response["choices"][0]["message"]["content"].strip()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--software-only", action="store_true",
                        help="Only process rows where L10 = 'IBM Software'")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print prompts without calling the API")
    args = parser.parse_args()

    if not EXCEL_PATH.exists():
        sys.exit(f"Error: Excel file not found at {EXCEL_PATH}")

    print(f"Loading {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    model = get_model(args.dry_run)

    total = 0
    skipped = 0
    updated = 0
    errors = 0

    # Row 1 is the header; data starts at row 2
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        l10  = str(row[COL_L10  - 1].value or "").strip()
        l15  = str(row[COL_L15  - 1].value or "").strip()
        l17  = str(row[COL_L17  - 1].value or "").strip()
        l20  = str(row[COL_L20  - 1].value or "").strip()
        l30  = str(row[COL_L30  - 1].value or "").strip()
        desc_cell = row[COL_DESC - 1]
        # MergedCell instances are read-only; skip them
        if desc_cell.__class__.__name__ == "MergedCell":
            skipped += 1
            continue
        existing  = str(desc_cell.value or "").strip()

        total += 1

        # Skip if already has a description
        if existing:
            skipped += 1
            continue

        # Skip Infrastructure rows if --software-only
        if args.software_only and l10 != "IBM Software":
            skipped += 1
            continue

        # Skip blank L30
        if not l30:
            skipped += 1
            continue

        # Skip known vague/internal entries
        if l30 in SKIP_VALUES:
            skipped += 1
            print(f"  SKIP  [{l30}]")
            continue

        prompt = build_prompt(l10, l15, l17, l20, l30)

        if args.dry_run:
            print(f"\n── DRY RUN ──────────────────────────────────────────")
            print(prompt)
            updated += 1
            continue

        print(f"  GEN   [{l30}] ...", end=" ", flush=True)
        try:
            description = generate_description(model, prompt)
            ws.cell(row=row_idx, column=COL_DESC).value = description
            wb.save(EXCEL_PATH)   # save after every row
            print(f"OK")
            updated += 1
        except Exception as e:
            print(f"ERROR: {e}")
            errors += 1
            time.sleep(2)  # brief pause on error before continuing

    print(f"\nDone. {total} rows total | {updated} updated | {skipped} skipped | {errors} errors")


if __name__ == "__main__":
    main()
